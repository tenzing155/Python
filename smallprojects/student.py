from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.geometry('700x600+400+100')
root.resizable(False,False)
root.configure(bg="#326273")


file = pathlib.Path('smallprojects\student.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['H6'] = "Full Name"
    sheet['I6'] = "Phone Number"
    sheet['J6'] = "Age"
    sheet['K6'] = "Gender"
    sheet['L6'] = "Address"
    sheet['M6'] = "Grade"
    sheet['N6'] = "Birthday"

    file.save('student.xlsx')

def submit():
    name = nameValue.get()
    phone = phoneValue.get()
    age = ageValue.get()
    gender = gender_combobox.get()
    address = addressValue.get()
    grade = gradeValue.get()
    birthday = birthValue.get()

    file = openpyxl.load_workbook('smallprojects\student.xlsx')
    sheet = file.active


    row = 7
    while sheet.cell(row=row, column=8).value is not None:
        row += 1

    sheet.cell(column=8, row=row, value=name)
    sheet.cell(column=9, row=row, value=phone)
    sheet.cell(column=10, row=row, value=age)
    sheet.cell(column=11, row=row, value=gender)
    sheet.cell(column=12, row=row, value=address)
    sheet.cell(column=13, row=row, value=grade)
    sheet.cell(column=14, row=row, value=birthday)

    file.save('smallprojects\student.xlsx')

    messagebox.showinfo('info', "Details Added!")
    clear()


 
def clear():
    nameValue.set('')
    phoneValue.set('')
    ageValue.set('')
    addressValue.set('')
    gradeValue.set('')
    birthValue.set('')

#heading
Label(root, text="Please Enter your details:", font="arial 13", bg="#326273", fg="#fff").place(x=270,y=20)

#label
Label(root,text='Name', font=23, bg="#326273", fg="#fff").place(x=50,y=100)
Label(root,text='Phone', font=23, bg="#326273", fg="#fff").place(x=50,y=150)
Label(root,text='Age', font=23, bg="#326273", fg="#fff").place(x=50,y=200)
Label(root,text='Gender', font=23, bg="#326273", fg="#fff").place(x=370,y=200)
Label(root,text='Address', font=23, bg="#326273", fg="#fff").place(x=50,y=250)
Label(root,text='Grade', font=23, bg="#326273", fg="#fff").place(x=50,y=300)
Label(root,text='Birthday', font=23, bg="#326273", fg="#fff").place(x=50,y=350)


#Entry
nameValue = StringVar()
phoneValue = StringVar()
ageValue = StringVar()
addressValue = StringVar()
gradeValue = StringVar()
birthValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
nameEntry.place(x=150,y=100)
phoneEntry = Entry(root,textvariable=phoneValue,width=45,bd=2,font=20)
phoneEntry.place(x=150,y=150)
ageEntry = Entry(root,textvariable=ageValue,width=15,bd=2,font=20)
ageEntry.place(x=150,y=200)
gender_combobox = Combobox(root,values=['Male','Female'],font="arial 14", state='r',width=14)
gender_combobox.place(x=450,y=200)
gender_combobox.set('Male')
addressEntry = Entry(root,textvariable=addressValue,width=25,bd=2,font=20)
addressEntry.place(x=150,y=250)
gradeEntry = Entry(root,textvariable=gradeValue,width=25,bd=2,font=20)
gradeEntry.place(x=150,y=300)
birthEntry = Entry(root,textvariable=birthValue,width=25,bd=2,font=20)
birthEntry.place(x=150,y=350)

#button
Button(root,text="Submit",bg="white", fg="#326273",width=15, height=2, command=submit).place(x=150,y=450)
Button(root,text="Clear",bg="white", fg="#326273",width=15, height=2, command=clear).place(x=350,y=450)




root.mainloop()